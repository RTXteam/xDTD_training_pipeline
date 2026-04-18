import sys
import os
import argparse
import json
import polars as pl
from openpyxl import load_workbook

## Import Personal Packages
pathlist = os.getcwd().split(os.path.sep)
ROOTindex = pathlist.index("xDTD_training_pipeline")
ROOTPath = os.path.sep.join([*pathlist[:(ROOTindex + 1)]])
sys.path.append(os.path.join(ROOTPath, 'scripts'))
import utils

EXTRACT_COLS = {
    'final normalized drug id': 'drug_id',
    'final normalized drug label': 'drug_name',
    'final normalized disease id': 'disease_id',
    'final normalized disease label': 'disease_name',
}

VALID_DRUG_CATEGORIES = {'biolink:SmallMolecule', 'biolink:Drug', 'biolink:ChemicalEntity'}
VALID_DISEASE_CATEGORIES = {'biolink:Disease', 'biolink:PhenotypicFeature'}


def read_xlsx_to_polars(path, logger):
    """Read an Excel sheet into a Polars DataFrame via openpyxl."""
    wb = load_workbook(path, read_only=True)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    data = [
        {h: (str(v) if v is not None else None) for h, v in zip(headers, row)}
        for row in ws.iter_rows(min_row=2, values_only=True)
    ]
    wb.close()
    logger.info(f"Read {len(data)} rows from {os.path.basename(path)}")
    schema = {h: pl.Utf8 for h in headers}
    return pl.DataFrame(data, schema=schema)


def normalize_pairs(df, biolink_version, logger):
    """Normalize drug/disease CURIEs via Node Norm API and attach category info.

    Returns a DataFrame with columns:
      drug_id, drug_name, drug_primary_category, drug_categories,
      disease_id, disease_name, disease_primary_category, disease_categories
    Rows are dropped if normalization fails or if the drug/disease categories
    do not intersect with VALID_DRUG_CATEGORIES / VALID_DISEASE_CATEGORIES.
    """
    all_curies = df['drug_id'].unique().to_list() + df['disease_id'].unique().to_list()
    logger.info(f"Batch-normalizing {len(set(all_curies))} unique CURIEs")
    utils.batch_normalize_curies(all_curies)

    rows = []
    dropped_norm = 0
    dropped_cat = 0
    for row in df.iter_rows(named=True):
        drug_info = utils.get_node_norm_info(row['drug_id'])
        disease_info = utils.get_node_norm_info(row['disease_id'])
        if drug_info is None or disease_info is None:
            dropped_norm += 1
            continue

        drug_cats = drug_info.get('types', [])
        disease_cats = disease_info.get('types', [])

        if not (set(drug_cats) & VALID_DRUG_CATEGORIES):
            dropped_cat += 1
            continue
        if not (set(disease_cats) & VALID_DISEASE_CATEGORIES):
            dropped_cat += 1
            continue

        rows.append({
            'drug_id': drug_info['preferred_curie'],
            'drug_name': drug_info.get('preferred_name') or row['drug_name'],
            'drug_primary_category': utils.get_primary_category(drug_cats, biolink_version),
            'drug_categories': json.dumps(drug_cats),
            'disease_id': disease_info['preferred_curie'],
            'disease_name': disease_info.get('preferred_name') or row['disease_name'],
            'disease_primary_category': utils.get_primary_category(disease_cats, biolink_version),
            'disease_categories': json.dumps(disease_cats),
        })

    logger.info(f"Dropped {dropped_norm} rows (normalization failed), {dropped_cat} rows (category mismatch)")
    result = pl.DataFrame(rows)
    before = result.height
    result = result.unique(subset=['drug_id', 'disease_id'])
    logger.info(f"Deduplicated: {before} → {result.height} unique pairs")
    return result


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description="Process ground truth indication/contraindication Excel files")
    parser.add_argument("--log_dir", type=str, help="Log folder path", default=os.path.join(ROOTPath, "log_folder"))
    parser.add_argument("--log_name", type=str, help="Log file name", default="step5_process_ground_truth_pairs.log")
    parser.add_argument("--indication_file", type=str, help="Path to indicationList.xlsx", required=True)
    parser.add_argument("--contraindication_file", type=str, help="Path to contraindicationList.xlsx", required=True)
    parser.add_argument("--drug_list", type=str, help="Path to drug_list.txt from step4", required=True)
    parser.add_argument("--disease_list", type=str, help="Path to disease_list.txt from step4", required=True)
    parser.add_argument("--graph_nodes", type=str, help="Path to filtered_graph_nodes_info.txt", required=True)
    parser.add_argument("--biolink_version", type=str, help="Biolink model version", default="4.2.0")
    parser.add_argument("--output_folder", type=str, help="Output folder", default=os.path.join(ROOTPath, "data", "ground_truth_pairs"))
    args = parser.parse_args()

    logger = utils.get_logger(os.path.join(args.log_dir, args.log_name))
    logger.info(args)
    os.makedirs(args.output_folder, exist_ok=True)

    ## ── Load existing drug/disease lists from step4 ─────────────────────
    existing_drug_df = pl.read_csv(args.drug_list, separator='\t')
    existing_disease_df = pl.read_csv(args.disease_list, separator='\t')
    existing_drug_ids = set(existing_drug_df['drug_id'].to_list())
    existing_disease_ids = set(existing_disease_df['disease_id'].to_list())
    logger.info(f"Loaded {len(existing_drug_ids)} drugs, {len(existing_disease_ids)} diseases from step4 lists")

    ## ── Load KG node IDs for augmentation filtering ─────────────────────
    kg_node_ids = set(pl.read_csv(args.graph_nodes, separator='\t')['id'].to_list())
    logger.info(f"KG contains {len(kg_node_ids)} nodes")

    ## ── Indication list → TP pairs ──────────────────────────────────────
    logger.info("=== Processing indication list (TP pairs) ===")
    ind_df = read_xlsx_to_polars(args.indication_file, logger)
    ind_df = ind_df.select([
        pl.col(orig).alias(new) for orig, new in EXTRACT_COLS.items()
    ]).drop_nulls()
    logger.info(f"{ind_df.height} rows after column selection and null removal")

    tp_pairs = normalize_pairs(ind_df, args.biolink_version, logger)

    ## ── Contraindication list → TN pairs ────────────────────────────────
    logger.info("=== Processing contraindication list (TN pairs) ===")
    contra_df = read_xlsx_to_polars(args.contraindication_file, logger)

    before = contra_df.height
    contra_df = contra_df.filter(
        (pl.col('is_allergen').cast(pl.Utf8).str.strip_chars().str.to_uppercase() != 'TRUE')
        & (pl.col('is_diagnostic_agent').cast(pl.Utf8).str.strip_chars().str.to_uppercase() != 'TRUE')
    )
    logger.info(f"Filtered allergens/diagnostic agents: {before} → {contra_df.height}")

    contra_df = contra_df.select([
        pl.col(orig).alias(new) for orig, new in EXTRACT_COLS.items()
    ]).drop_nulls()
    logger.info(f"{contra_df.height} rows after column selection and null removal")

    tn_pairs = normalize_pairs(contra_df, args.biolink_version, logger)

    ## ── Augment drug/disease lists with new entries from ground truth ────
    logger.info("=== Augmenting drug/disease lists ===")
    all_pairs = pl.concat([tp_pairs, tn_pairs])

    new_drug_rows = []
    for drug_id in all_pairs['drug_id'].unique().to_list():
        if drug_id not in existing_drug_ids and drug_id in kg_node_ids:
            info = utils.get_node_norm_info(drug_id)
            if info is not None:
                cats = info.get('types', [])
                new_drug_rows.append({
                    'drug_id': drug_id,
                    'drug_name': info.get('preferred_name', ''),
                    'primary_category': utils.get_primary_category(cats, args.biolink_version),
                    'categories': json.dumps(cats),
                })

    new_disease_rows = []
    for disease_id in all_pairs['disease_id'].unique().to_list():
        if disease_id not in existing_disease_ids and disease_id in kg_node_ids:
            info = utils.get_node_norm_info(disease_id)
            if info is not None:
                cats = info.get('types', [])
                new_disease_rows.append({
                    'disease_id': disease_id,
                    'disease_name': info.get('preferred_name', ''),
                    'primary_category': utils.get_primary_category(cats, args.biolink_version),
                    'categories': json.dumps(cats),
                })

    logger.info(f"New drugs to add (present in KG): {len(new_drug_rows)}")
    logger.info(f"New diseases to add (present in KG): {len(new_disease_rows)}")

    if new_drug_rows:
        updated_drug_df = pl.concat([existing_drug_df, pl.DataFrame(new_drug_rows)])
        updated_drug_df.write_csv(args.drug_list, separator='\t')
        logger.info(f"Updated drug_list: {existing_drug_df.height} → {updated_drug_df.height}")
    else:
        logger.info("Drug list unchanged")

    if new_disease_rows:
        updated_disease_df = pl.concat([existing_disease_df, pl.DataFrame(new_disease_rows)])
        updated_disease_df.write_csv(args.disease_list, separator='\t')
        logger.info(f"Updated disease_list: {existing_disease_df.height} → {updated_disease_df.height}")
    else:
        logger.info("Disease list unchanged")

    ## ── Filter pairs to only those with both drug and disease in KG ──────
    logger.info("=== Filtering pairs by KG presence ===")
    before_tp = tp_pairs.height
    tp_pairs = tp_pairs.filter(
        pl.col('drug_id').is_in(kg_node_ids) & pl.col('disease_id').is_in(kg_node_ids)
    )
    logger.info(f"TP pairs filtered by KG: {before_tp} → {tp_pairs.height}")
    tp_path = os.path.join(args.output_folder, 'tp_pairs.txt')
    tp_pairs.write_csv(tp_path, separator='\t')
    logger.info(f"Saved {tp_pairs.height} TP pairs to {tp_path}")

    before_tn = tn_pairs.height
    tn_pairs = tn_pairs.filter(
        pl.col('drug_id').is_in(kg_node_ids) & pl.col('disease_id').is_in(kg_node_ids)
    )
    logger.info(f"TN pairs filtered by KG: {before_tn} → {tn_pairs.height}")
    tn_path = os.path.join(args.output_folder, 'tn_pairs.txt')
    tn_pairs.write_csv(tn_path, separator='\t')
    logger.info(f"Saved {tn_pairs.height} TN pairs to {tn_path}")

